# report.py — генерация Excel / CSV отчётов
# Структура Excel:
#   Лист "Детализация"        — сводная таблица всех сайтов
#   Лист "Статистика"         — агрегированная статистика
#   Лист "Аудит: domain.ru"   — детальный разбор по критериям для каждого сайта

import io
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import RUSSIAN_ANALYTICS

# ── Цвета ────────────────────────────────────────────────────────────────────
RISK_COLORS = {
    "ВЫСОКИЙ": "FFD7D7",
    "СРЕДНИЙ":  "FFF3CD",
    "НИЗКИЙ":   "D4EDDA",
}

# Цвета строк детального листа по категории нарушения
CAT_COLORS = {
    "ok":      ("E8F5E9", "388E3C"),   # (fill, font)  зелёный
    "warning": ("FFF8E1", "F57F17"),   # жёлтый
    "error":   ("FFEBEE", "C62828"),   # красный
    "info":    ("F3F3F3", "555555"),   # серый
    "header":  ("37474F", "FFFFFF"),   # тёмный заголовок
    "section": ("E3F2FD", "1565C0"),   # синий раздел
}

# Штрафы: категория → (статья КоАП, размер штрафа)
FINES = {
    "foreign": ("ч.9 ст.13.11 КоАП РФ", "до 500 000 ₽"),
    "cookie":  ("ч.3 ст.13.11 КоАП РФ", "до 300 000 ₽"),
    "policy":  ("ч.2 ст.13.11 КоАП РФ", "до 150 000 ₽"),
    "forms":   ("ч.2 ст.13.11 КоАП РФ", "до 100 000 ₽"),
}

# Сводные столбцы
COLUMNS = {
    "pages_count":             "Страниц проверено",
    "pages_checked":           "Проверенные страницы",
    "domain":                  "Домен",
    "accessible":              "Доступен",
    "http_status":             "HTTP",
    "policy_found":            "Политика ПД",
    "policy_in_footer":        "Политика в футере",
    "policy_url":              "URL политики",
    "policy_text_len":         "Длина политики",
    "policy_is_pdf":           "Политика в PDF",
    "sections_missing":        "Отсутствующие разделы политики",
    "operator_found":          "Реквизиты оператора",
    "has_cookie_banner":       "Cookie-баннер",
    "has_decline_button":      "Кнопка «Отказаться»",
    "checked_by_default":      "Галочка предустановлена",
    "analytics_systems":       "Системы аналитики",
    "foreign_resources":       "Иностранные ресурсы",
    "pd_forms_count":          "Форм с ПД",
    "pd_fields":               "Поля форм",
    "forms_have_warning":      "Предупреждение у форм",
    "consent_level":           "Согласие на ПД",
    "any_text_consent":        "Текстовое согласие (без чекбокса)",
    "missing_requisites":      "Отсутствующие реквизиты",
    "consent_violations":      "Нарушения в согласии",
    "playwright_used":         "Playwright",
    "banner_visible":          "Баннер виден (Playwright)",
    "trackers_before_consent": "Трекеры до согласия",
    "accept_btn_found":        "Кнопка «Принять» (Playwright)",
    "decline_btn_found":       "Кнопка «Отказаться» (Playwright)",
    "risk":                    "Оценка риска",
    "violations":              "Нарушения и штрафы",
    "check_time_sec":          "Время (сек)",
    "error":                   "Комментарий",
}


# ─────────────────────────────────────────────────────────────────────────────
# Вспомогательные функции
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(v) -> str:
    if isinstance(v, bool):
        return "Да" if v else "Нет"
    if isinstance(v, list):
        return "; ".join(str(i) for i in v) if v else "—"
    return str(v) if v not in (None, "") else "—"


def _row(r: dict) -> dict:
    return {header: _fmt(r.get(key, "")) for key, header in COLUMNS.items()}


def _safe_sheet_name(domain: str, existing: set) -> str:
    """Безопасное имя листа: макс 31 символ, без спецсимволов, уникальное."""
    # Excel запрещает символы: \ / * ? [ ] :
    name = re.sub(r"[\\/*?\[\]:]", "_", domain)[:24]
    name = f"Аудит {name}"[:31]
    base, n = name, 2
    while name in existing:
        name = f"{base[:28]}_{n}"
        n += 1
    existing.add(name)
    return name


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(color: str = "000000", bold: bool = False, size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border_bottom() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(bottom=side)


def _set_col_width(ws, col_idx: int, width: int):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─────────────────────────────────────────────────────────────────────────────
# Построение детального листа одного сайта
# ─────────────────────────────────────────────────────────────────────────────

def _build_site_rows(r: dict) -> list[dict]:
    """
    Формирует список строк для детального листа сайта.
    Каждая строка: {section, status_icon, description, article, fine, cat}
    """
    rows = []

    def ok(section, desc):
        rows.append({"section": section, "icon": "✓", "desc": desc,
                     "article": "", "fine": "", "cat": "ok"})

    def warn(section, desc, cat="warning", article="", fine=""):
        rows.append({"section": section, "icon": "⚠", "desc": desc,
                     "article": article, "fine": fine, "cat": cat})

    def err(section, desc, cat="error", article="", fine=""):
        rows.append({"section": section, "icon": "✗", "desc": desc,
                     "article": article, "fine": fine, "cat": cat})

    def info(section, desc):
        rows.append({"section": section, "icon": "ℹ", "desc": desc,
                     "article": "", "fine": "", "cat": "info"})

    # ── 1. Политика обработки ПД ─────────────────────────────────────────────
    S = "1. Политика обработки ПД"
    if not r.get("policy_found"):
        err(S, "Политика обработки персональных данных не найдена на сайте",
            article=FINES["policy"][0], fine=FINES["policy"][1])
    else:
        url = r.get("policy_url", "")
        tlen = r.get("policy_text_len", 0)
        ok(S, f"Политика найдена: {url}  ({tlen:,} символов)")

        if not r.get("policy_in_footer"):
            warn(S, "Ссылка на политику отсутствует в футере сайта — "
                 "РКН требует доступность не более чем за 2 клика",
                 article=FINES["policy"][0], fine=FINES["policy"][1])
        else:
            ok(S, "Ссылка на политику размещена в футере сайта")

        if r.get("policy_is_pdf"):
            warn(S, "Политика размещена в формате PDF — "
                 "АС МПДн Роскомнадзора не парсит PDF, может не засчитать документ",
                 cat="warning")
        else:
            ok(S, "Политика в формате HTML — корректно читается АС МПДн")

        missing_sec = r.get("sections_missing") or []
        found_sec   = r.get("sections_found")   or []
        if missing_sec:
            err(S, f"Отсутствуют обязательные разделы по ст.14 152-ФЗ "
                f"({len(missing_sec)} из 11): {'; '.join(missing_sec)}",
                article=FINES["policy"][0], fine=FINES["policy"][1])
        else:
            ok(S, f"Все {len(found_sec)} обязательных разделов по ст.14 152-ФЗ присутствуют")

        if not r.get("operator_found"):
            warn(S, "Реквизиты оператора (ИНН, ОГРН, юр. адрес) не найдены в тексте политики",
                 article=FINES["policy"][0], fine=FINES["policy"][1])
        else:
            ok(S, "Реквизиты оператора (ИНН/ОГРН/адрес) найдены в тексте политики")

    # ── 2. Cookie-баннер и метрика ────────────────────────────────────────────
    S = "2. Cookie-баннер и метрика"
    systems = r.get("analytics_systems") or []
    foreign = r.get("foreign_resources") or []

    if systems:
        info(S, f"Обнаружены системы аналитики: {', '.join(systems)}")
        if not r.get("has_cookie_banner"):
            err(S, "Cookie-баннер отсутствует при наличии трекеров — "
                "требуется получить явное согласие пользователя до загрузки трекеров",
                article=FINES["cookie"][0], fine=FINES["cookie"][1])
        else:
            ok(S, "Cookie-баннер присутствует")
            if not r.get("has_decline_button"):
                err(S, "В cookie-баннере отсутствует кнопка «Отказаться» / "
                    "«Только необходимые» — пользователь должен иметь возможность отказаться",
                    article=FINES["cookie"][0], fine=FINES["cookie"][1])
            else:
                ok(S, "Кнопка отказа от cookies присутствует в баннере")

        if r.get("checked_by_default"):
            err(S, "Галочка согласия на cookie предустановлена по умолчанию — "
                "принудительное согласие прямо запрещено",
                article=FINES["cookie"][0], fine=FINES["cookie"][1])
        else:
            ok(S, "Галочки согласия не предустановлены — норма соблюдена")
    else:
        ok(S, "Системы аналитики не обнаружены — cookie-согласие не требуется")

    # ── 3. Иностранные ресурсы ────────────────────────────────────────────────
    # foreign_resources включает CDN/виджеты + иностранную аналитику (GA, Meta,
    # Hotjar, SimilarWeb). Яндекс.Метрика / LiveInternet / Roistat — нет.
    S = "3. Иностранные ресурсы (трансграничная передача)"
    if foreign:
        err(S, f"Обнаружены иностранные ресурсы / аналитика без уведомления РКН — "
            f"каждое обращение к ним передаёт IP-адрес пользователя за рубеж: "
            f"{', '.join(foreign)}",
            article=FINES["foreign"][0], fine=FINES["foreign"][1])
        info(S, "Решение: заменить на российские аналоги или захостить локально; "
             "подать уведомление в РКН о трансграничной передаче. "
             "Яндекс.Метрика трансграничной передачей не является.")
    else:
        ok(S, "Иностранные ресурсы и иностранная аналитика не обнаружены — "
           "трансграничная передача отсутствует "
           "(российская аналитика, в т.ч. Яндекс.Метрика, сюда не относится)")

    # ── 4. Playwright: трекеры до согласия ───────────────────────────────────
    if r.get("playwright_used"):
        S = "4. Проверка поведения (Playwright)"
        trackers_before = r.get("trackers_before_consent") or []
        if trackers_before:
            only_russian = all(t in RUSSIAN_ANALYTICS for t in trackers_before)
            fine_key = "cookie" if only_russian else "foreign"
            err(S, f"Трекеры загружаются ДО получения согласия пользователя: "
                f"{', '.join(trackers_before)} — "
                "нарушение: аналитика должна запускаться только после клика «Принять»"
                + ("; российская аналитика — не трансграничная передача, "
                   "но cookie-согласие всё равно требуется" if only_russian else ""),
                article=FINES[fine_key][0], fine=FINES[fine_key][1])
        else:
            ok(S, "Трекеры не загружаются до получения согласия — норма соблюдена")

        if not r.get("banner_visible"):
            warn(S, "«Спрятанный баннер»: cookie-баннер не отображается при первом посещении "
                 "(обнаружено Playwright в режиме инкогнито) — "
                 "АС МПДн работает аналогично и не увидит баннер",
                 article=FINES["cookie"][0], fine=FINES["cookie"][1])
        else:
            ok(S, "Cookie-баннер отображается при первом посещении в чистом браузере")
    else:
        S = "4. Проверка поведения (Playwright)"
        info(S, "Playwright не использовался — включите «Углублённую проверку» "
             "для обнаружения трекеров до согласия и спрятанных баннеров")

    # ── 5. Формы сбора ПД ────────────────────────────────────────────────────
    S = "5. Формы сбора персональных данных"
    forms_count = r.get("pd_forms_count") or 0
    if forms_count == 0:
        ok(S, "Формы сбора персональных данных не обнаружены")
    else:
        info(S, f"Обнаружено форм с полями ПД: {forms_count}")
        fields = r.get("pd_fields") or []
        if fields:
            # Выводим уникальные короткие названия полей
            short = []
            for f in fields:
                m = re.search(r"placeholder=([^\]]+)", f)
                if m:
                    short.append(m.group(1))
                else:
                    m2 = re.search(r"name=([^,\]]+)", f)
                    if m2:
                        short.append(m2.group(1))
            info(S, f"Поля ПД в формах: {', '.join(dict.fromkeys(short)) or '—'}")

        if r.get("any_text_consent"):
            warn(S, "Согласие оформлено только текстом «Нажимая Отправить...» — "
                 "152-ФЗ требует отдельный чекбокс с явным согласием, "
                 "текстовая конструкция не является надлежащим согласием",
                 article=FINES["forms"][0], fine=FINES["forms"][1])

    # ── 6. Согласие на обработку ПД ──────────────────────────────────────────
    S = "6. Согласие на обработку персональных данных"
    consent = r.get("consent_level") or "не проверено"
    if forms_count == 0:
        info(S, "Форм с ПД нет — проверка согласия не требуется")
    elif consent == "полное":
        ok(S, "Согласие оформлено корректно: отдельный чекбокс + все обязательные "
           "реквизиты по 152-ФЗ присутствуют")
    elif consent == "частичное":
        missing_req = r.get("missing_requisites") or []
        warn(S, f"Согласие частичное: чекбокс есть, но отсутствуют обязательные "
             f"реквизиты по 152-ФЗ: {', '.join(missing_req) or '—'}",
             article=FINES["forms"][0], fine=FINES["forms"][1])
    elif consent == "отсутствует":
        err(S, "Согласие на обработку ПД полностью отсутствует — "
            "нет чекбокса согласия рядом с формой",
            article=FINES["forms"][0], fine=FINES["forms"][1])
    elif consent and "текстовое" in consent:
        missing_req = r.get("missing_requisites") or []
        err(S, f"Согласие только текстом — не соответствует 152-ФЗ. "
            f"Нужен отдельный чекбокс. "
            f"Отсутствуют реквизиты: {', '.join(missing_req) or '—'}",
            article=FINES["forms"][0], fine=FINES["forms"][1])
    elif consent == "нарушения":
        violations = r.get("consent_violations") or []
        for v in violations:
            err(S, f"Нарушение: {v}",
                article=FINES["forms"][0], fine=FINES["forms"][1])
    else:
        info(S, f"Статус согласия: {consent}")

    return rows


def _write_site_sheet(wb, r: dict, sheet_name: str):
    """Записывает детальный лист аудита одного сайта в workbook."""
    ws = wb.create_sheet(title=sheet_name)

    # ── Заголовок ─────────────────────────────────────────────────────────
    domain = r.get("domain", r.get("url", "—"))
    risk   = r.get("risk", "—")
    risk_color = "FFD7D7" if "ВЫСОКИЙ" in risk else \
                 "FFF3CD" if "СРЕДНИЙ"  in risk else "D4EDDA"

    ws.merge_cells("A1:F1")
    ws["A1"].value     = f"Аудит сайта: {domain}"
    ws["A1"].font      = _font("FFFFFF", bold=True, size=13)
    ws["A1"].fill      = _fill("263238")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=False)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Оценка риска: {risk}   |   Время проверки: {r.get('check_time_sec', 0)} сек   |   URL: {r.get('url', '—')}"
    ws["A2"].font      = _font("333333", bold=False, size=10)
    ws["A2"].fill      = _fill(risk_color)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # Пустая строка-разделитель
    ws.row_dimensions[3].height = 6

    # ── Шапка таблицы ─────────────────────────────────────────────────────
    headers = ["Критерий / Раздел", "Ст.", "Описание результата проверки",
               "Статья КоАП РФ", "Вероятный штраф", ""]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = _font("FFFFFF", bold=True, size=10)
        c.fill      = _fill("37474F")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[4].height = 22

    # ── Строки данных ─────────────────────────────────────────────────────
    site_rows = _build_site_rows(r)
    current_section = None
    row_num = 5

    for d in site_rows:
        # Если сменился раздел — вставляем заголовок секции
        if d["section"] != current_section:
            current_section = d["section"]
            ws.merge_cells(f"A{row_num}:F{row_num}")
            c = ws.cell(row=row_num, column=1, value=current_section)
            c.font      = _font("1565C0", bold=True, size=10)
            c.fill      = _fill("E3F2FD")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Цвет строки
        bg_color, fg_color = CAT_COLORS.get(d["cat"], CAT_COLORS["info"])
        row_fill = _fill(bg_color)

        # Иконка статуса
        c1 = ws.cell(row=row_num, column=1, value="")
        c1.fill = row_fill

        c2 = ws.cell(row=row_num, column=2, value=d["icon"])
        c2.font      = _font(fg_color, bold=True, size=11)
        c2.fill      = row_fill
        c2.alignment = Alignment(horizontal="center", vertical="top")

        c3 = ws.cell(row=row_num, column=3, value=d["desc"])
        c3.font      = _font("333333", size=10)
        c3.fill      = row_fill
        c3.alignment = Alignment(wrap_text=True, vertical="top", indent=1)

        c4 = ws.cell(row=row_num, column=4, value=d.get("article") or "")
        c4.font      = _font("666666", size=9)
        c4.fill      = row_fill
        c4.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

        c5 = ws.cell(row=row_num, column=5, value=d.get("fine") or "")
        c5.font      = _font(fg_color if d["cat"] == "error" else "666666",
                             bold=(d["cat"] == "error"), size=10)
        c5.fill      = row_fill
        c5.alignment = Alignment(horizontal="right", vertical="top")

        # Пустая 6-я для красоты
        ws.cell(row=row_num, column=6).fill = row_fill

        # Нижняя граница строки
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = _border_bottom()

        ws.row_dimensions[row_num].height = 32
        row_num += 1

    # ── Итоговая строка: суммарный штраф ──────────────────────────────────
    row_num += 1
    violations = r.get("violations") or []
    if violations:
        ws.merge_cells(f"A{row_num}:B{row_num}")
        ws.cell(row=row_num, column=1, value="ИТОГО нарушений:").font = \
            _font("C62828", bold=True, size=10)
        ws.cell(row=row_num, column=1).fill = _fill("FFEBEE")

        ws.merge_cells(f"C{row_num}:F{row_num}")
        ws.cell(row=row_num, column=3,
                value=f"{len(violations)} нарушений | {'; '.join(violations[:3])}"
                      f"{'...' if len(violations) > 3 else ''}").font = \
            _font("C62828", size=10)
        ws.cell(row=row_num, column=3).fill  = _fill("FFEBEE")
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 40

    # ── Ширина столбцов ───────────────────────────────────────────────────
    _set_col_width(ws, 1, 3)    # пустой отступ
    _set_col_width(ws, 2, 4)    # иконка
    _set_col_width(ws, 3, 70)   # описание
    _set_col_width(ws, 4, 26)   # статья
    _set_col_width(ws, 5, 18)   # штраф
    _set_col_width(ws, 6, 3)    # отступ

    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────────────────────────────────────
# Сводная статистика
# ─────────────────────────────────────────────────────────────────────────────

def _stats(results: list) -> pd.DataFrame:
    total = len(results)
    if not total:
        return pd.DataFrame()

    w_pol  = sum(1 for r in results if r.get("policy_found"))
    w_foot = sum(1 for r in results if r.get("policy_in_footer"))
    w_ban  = sum(1 for r in results if r.get("has_cookie_banner"))
    w_dec  = sum(1 for r in results if r.get("has_decline_button"))
    w_for  = sum(1 for r in results if r.get("foreign_resources"))
    w_oper = sum(1 for r in results if r.get("operator_found"))
    high   = sum(1 for r in results if "ВЫСОКИЙ" in r.get("risk", ""))
    med    = sum(1 for r in results if "СРЕДНИЙ"  in r.get("risk", ""))
    low    = sum(1 for r in results if "НИЗКИЙ"   in r.get("risk", ""))
    pw_cnt = sum(1 for r in results if r.get("playwright_used"))
    pct    = lambda n: f"{n/total*100:.0f}%" if total else "0%"

    viols = sorted([
        ("Нет политики ПД",                     total - w_pol),
        ("Политика не в футере",                total - w_foot),
        ("Нет реквизитов оператора",            total - w_oper),
        ("Нет cookie-баннера",                  total - w_ban),
        ("Нет кнопки «Отказаться»",             total - w_dec),
        ("Иностранные ресурсы (трансграничка)",  w_for),
        ("Нет согласия на ПД",
         sum(1 for r in results if r.get("consent_level") == "отсутствует")),
        ("Трекеры грузятся до согласия",
         sum(1 for r in results if r.get("trackers_before_consent"))),
    ], key=lambda x: -x[1])

    rows = [
        ["Всего сайтов проверено",    total,  ""],
        ["С политикой ПД",            w_pol,  pct(w_pol)],
        ["Политика в футере",         w_foot, pct(w_foot)],
        ["С реквизитами оператора",   w_oper, pct(w_oper)],
        ["С cookie-баннером",         w_ban,  pct(w_ban)],
        ["С кнопкой «Отказаться»",    w_dec,  pct(w_dec)],
        ["Иностранные ресурсы",       w_for,  pct(w_for)],
        ["Проверено через Playwright", pw_cnt, pct(pw_cnt)],
        ["", "", ""],
        ["🔴 Высокий риск",          high,   pct(high)],
        ["🟡 Средний риск",          med,    pct(med)],
        ["🟢 Низкий риск",           low,    pct(low)],
        ["", "", ""],
        ["ТОП НАРУШЕНИЙ", "Кол-во", "%"],
    ] + [[n, c, pct(c)] for n, c in viols if c > 0]

    return pd.DataFrame(rows, columns=["Показатель", "Значение", "%"])


# ─────────────────────────────────────────────────────────────────────────────
# Главная функция генерации Excel
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(results: list) -> bytes:
    df_det  = pd.DataFrame([_row(r) for r in results],
                           columns=list(COLUMNS.values()))
    df_stat = _stats(results)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_det.to_excel(w,  sheet_name="Детализация", index=False)
        df_stat.to_excel(w, sheet_name="Статистика",  index=False)
    buf.seek(0)

    wb    = load_workbook(buf)
    hfill = PatternFill("solid", fgColor="D9D9D9")

    # Оформление листа Детализация
    for sname in ("Детализация", "Статистика"):
        ws = wb[sname]
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = hfill
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

    ws = wb["Детализация"]
    risk_col = next((i for i, c in enumerate(ws[1], 1)
                     if c.value == "Оценка риска"), None)
    for row in ws.iter_rows(min_row=2):
        if risk_col:
            rv = str(row[risk_col - 1].value or "")
            for key, color in RISK_COLORS.items():
                if key in rv:
                    fill = PatternFill("solid", fgColor=color)
                    for cell in row:
                        cell.fill = fill
                    break

    for sname in ("Детализация", "Статистика"):
        ws = wb[sname]
        for col in ws.columns:
            w = min(max(len(str(c.value or "")) for c in col) + 2, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = w
        ws.freeze_panes = "A2"

    # ── Листы для каждого сайта ───────────────────────────────────────────
    used_names: set = {"Детализация", "Статистика"}
    for r in results:
        domain     = r.get("domain") or r.get("url", "site")
        sheet_name = _safe_sheet_name(domain, used_names)
        _write_site_sheet(wb, r, sheet_name)

    # Ставим Детализацию первой
    wb.active = wb["Детализация"]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_csv(results: list) -> bytes:
    df = pd.DataFrame([_row(r) for r in results],
                      columns=list(COLUMNS.values()))
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
